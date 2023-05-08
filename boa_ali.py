#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Author: Daniel Sahlgren 2022, WSP

import emme_functions.emme as emme
import emme_functions.emme_functions
import emme_functions.emme_functions as ef
import emme_functions.auxiliary_functions as aux
import pandas as pd
import inro
import os
import pickle
from emme_functions.emme_functions import line_link_list as link_list
import inro.modeller as _m
from inro.emme.desktop import app as _app
import inro.emme.database.emmebank as _emmebank
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# R:/7055/10350700/5_Berakningar/Sampers/Rigg/Person2017_255_Nulage_230223
name_of_project = "Sthlm_Oslo_255_JA"
# project = 'D:/10350700_255_AE_DS/UA2_230504/E444bank//%s//%s//'
# project = 'R:/7055/10350700/5_Berakningar/Sampers/Rigg/Person2017_255_Nulage_230223_Resultatriggning/E443bank/%s/%s/'
project = 'R:/7055/10350700/5_Berakningar/Sampers/UA2_230505/E444bank/%s/%s/'
regional_bases =["Palt", "Samm", "Skane", "Sydost", "Vast"]
# regional_bases =["Samm"]
# regional_bases =["Palt"]
# my_desktop = _app.start_dedicated(project='R:/7055/10350700/5_Berakningar/Sampers/Rigg/Person2017_255_Nulage_230223_Resultatriggning/E443bank/UA/NB/Jvg/Jvg.emp', visible=False, user_initials="ds")
my_desktop = _app.start_dedicated(project='R:/7055/10350700/5_Berakningar/Sampers/UA2_230505/E444bank/UA/NB/Jvg/Jvg.emp', visible=False, user_initials="ds")
my_modeller = _m.Modeller(my_desktop)
print(my_desktop.version)


stations_sel = pd.read_csv("indata/stations_ali_boa.txt", encoding="utf-8")["station"].values.tolist()
print(stations_sel)
#
stations_sel = list(stations_sel)
scen_id = 1101
# decide if UA and/or JA sjould be analyzed
# scenario_ja_ua = ["JA", "UA"]
# scenario_ja_ua = ["UA"]
scenario_ja_ua = ["JA"]


nodes = {"JA": {}, "UA": {}}
year_factor = 365
# year_factor = 1

modes = ["i","j","k"]
#loop over regional models


for ja_ua in scenario_ja_ua:

    with _emmebank.Emmebank(project %(ja_ua,"NB") + "//Jvg////emmebank") as eb:
        node_names = {}

        scen = eb.scenario(scen_id)
        my_network = scen.get_network()

        my_network.create_attribute("NODE", "transfers_boa", default_value=0)
        my_network.create_attribute("NODE", "transfers_ali", default_value=0)
        my_network.create_attribute("NODE", "transit_boardings", default_value=0)
        my_network.create_attribute("TRANSIT_SEGMENT", "transit_alightings", default_value=0)
        my_network.create_attribute("NODE", "transit_alightings", default_value=0)

        # nodes[ja_ua] = {"nationell": {}}


        for line in my_network.transit_lines():
            if line.mode.id in modes:
                segments = line.segments(True)
                seg = next(segments)
                for next_seg in segments:
                    next_seg.transit_alightings = seg.transit_volume - next_seg.transit_volume + next_seg.transit_boardings
                    seg = next_seg

        for node in my_network.nodes():
            if node.is_centroid == False:

                ca_ali_i = 0

                for seg in node.outgoing_segments(include_hidden=True):
                    if seg.line.mode.id in modes:
                        ca_ali_i += seg.transit_alightings

                node.transfers_ali = ca_ali_i - node.final_alightings
                node.transit_alightings = ca_ali_i

                ca_board_i = 0
                for seg in node.outgoing_segments():
                    if seg.line.mode.id in modes:
                        ca_board_i += seg.transit_boardings

                node.transfers_boa = ca_board_i - node.initial_boardings
                node.transit_boardings = ca_board_i

                nodes[ja_ua][node.id] = {}
                # # nodes[ja_ua][node.id]["base"] = "nationell"



                nodes[ja_ua][node.id]["boa_tot"] = ca_board_i #TODO set back to this
                nodes[ja_ua][node.id]["boa_transfers"] = node.transfers_boa
                nodes[ja_ua][node.id]["boa_initial"] = node.initial_boardings
                nodes[ja_ua][node.id]["ali_tot"] = ca_ali_i
                nodes[ja_ua][node.id]["ali_transfers"] = node.transfers_ali
                nodes[ja_ua][node.id]["ali_final"] = node.final_alightings
                nodes[ja_ua][node.id]["ali_boa_tot"] = (ca_board_i + ca_ali_i)
                nodes[ja_ua][node.id]["station"] = node["#station"]
                if any(node["#station"] == item for item in stations_sel):
                    nodes[ja_ua][node.id]["station_selection"] = True
                    nodes[ja_ua][node.id]["station_order"] = stations_sel.index(node["#station"])



            # if ja_ua=="UA": # get station names from UA (needs to be one of JA or UA in national base (stations not imported to regional bases)
            #     node_names[node.id] = node["#station"]

    for regional_base in regional_bases:
        # project_file = project %(ja_ua, "RB") + regional_base + "//Koll//Koll.emp"

        with _emmebank.Emmebank(project %(ja_ua,"RB")  + regional_base + "//Koll//emmebank") as eb:

            scen = eb.scenario(scen_id)
            my_network = scen.get_network()
            my_network.create_attribute("NODE", "transfers_boa", default_value=0)
            my_network.create_attribute("NODE", "transfers_ali", default_value=0)
            my_network.create_attribute("NODE", "transit_boardings", default_value=0)
            my_network.create_attribute("TRANSIT_SEGMENT", "transit_alightings", default_value=0)
            my_network.create_attribute("NODE", "transit_alightings", default_value=0)

            for line in my_network.transit_lines():
                if line.mode.id in modes:
                    segments = line.segments(True)
                    seg = next(segments)
                    for next_seg in segments:
                        next_seg.transit_alightings = seg.transit_volume - next_seg.transit_volume + next_seg.transit_boardings
                        seg = next_seg

            for node in my_network.nodes():
                if node.is_centroid == False:
                    # calculate total alightings and transfer alightings
                    ca_ali_i = 0
                    for seg in node.outgoing_segments(include_hidden=True):
                        if seg.line.mode.id in modes:
                            ca_ali_i += seg.transit_alightings
                    node.transfers_ali = ca_ali_i - node.final_alightings
                    node.transit_alightings = ca_ali_i
                    # calculate total boardings and transfer boardings

                    ca_board_i = 0
                    for seg in node.outgoing_segments():
                        if seg.line.mode.id in modes:
                            ca_board_i += seg.transit_boardings
                    node.transfers_boa = ca_board_i - node.initial_boardings
                    node.transit_boardings = ca_board_i

            # for node in my_network.nodes():
            #     if node.id in nodes[ja_ua]: # check if node is in dictionary. If not this means that the node does not exist in national base and thus should be skipped
                    if node.id in nodes[ja_ua]: #there are more nodes in regional base than in national base
                        # nodes[ja_ua][node.id]["base"] = regional_base
                        nodes[ja_ua][node.id]["boa_tot"] += ca_board_i
                        nodes[ja_ua][node.id]["boa_transfers"] += node.transfers_boa
                        nodes[ja_ua][node.id]["boa_initial"] += node.initial_boardings
                        nodes[ja_ua][node.id]["ali_tot"] += ca_ali_i
                        nodes[ja_ua][node.id]["ali_transfers"] += node.transfers_ali
                        nodes[ja_ua][node.id]["ali_final"] += node.final_alightings
                        nodes[ja_ua][node.id]["ali_boa_tot"] += (ca_board_i + ca_ali_i)

                else:
                # except KeyError: # there are nodes in regional bases that dont exist in national. skip these
                    pass


        # convert to year numbers

    for node in nodes[ja_ua]:
        for key_, val_ in nodes[ja_ua][node].items():

            if key_ in ["station", "station_selection", "station_order"]:
                # print("passed")
                pass

            else:
                nodes[ja_ua][node][key_] *= year_factor

summary_ja = pd.DataFrame.from_dict(nodes["JA"], orient="index")
summary_ja["scenario"] = "JA"

summary_ua = pd.DataFrame.from_dict(nodes["UA"], orient="index")
summary_ua["scenario"] = "UA"

# summarize_nodes = {"JA", "UA"}
# for node, station in node_names.items():
#     if station !="":
#         summarize_nodes[station] = {"boa": 0, "ali": 0}
#         summarize_nodes[station] +=


summary = pd.concat([summary_ja, summary_ua])
summary.index.name='node_id'


summary_filtered = summary.dropna(subset=['station_selection'])
# summary_filtered = summary_filtered
print(summary_filtered)
name = "boa_ali_%s" % name_of_project
csv_name = name + ".csv"
# summary.to_csv(csv_name, sep=";", encoding="utf-8")
# emme_functions.auxiliary_functions.addUTF8Bom(csv_name)

import datetime
today = datetime.date.today()
info = {"project name": [name_of_project, ],"date": [datetime.date.today(), ],"catalogue": [project, ],"year factor": [year_factor,] ,"file created by": ["Daniel Sahlgren", ]} # collect info to save to info excel sheet
info =pd.DataFrame.from_dict(info, orient="index")
# info= pd.DataFrame.from_dict(info)
#create new df for storing information in info sheet in excel



#create Excel
wb = Workbook()
# ws = wb.active

# write all data to excelsheet



# Create a pivot table
ws_ali_boa = wb.create_sheet('ali_boa')
for r in dataframe_to_rows(summary, index=True, header=True):
    ws_ali_boa.append(r)


ws_NBB_stations = wb.create_sheet('ali_boa_stations')

# for r in dataframe_to_rows(summary_filtered, index=False, header=True):
#     ws_NBB_stations.append(r)

ws_info = wb.create_sheet('info')
# summary_filtered.sort_values("station_order", inplace=True)
pvt = summary_filtered.pivot_table(index=["station_order", "station"], aggfunc='sum', values="ali_boa_tot", columns=["scenario"]).reset_index(level=0, drop=True)
# pvt_sorted = summary_filtered.pivot_table(index='station', aggfunc='sum', values="ali_boa_tot", columns=["scenario"])
for r in dataframe_to_rows(pvt, index=True, header=True):
    ws_NBB_stations.append(r)

for r in dataframe_to_rows(info, index=True, header=True):
    ws_info.append(r)
    # ws_info.append("year factor " + year_factor)

if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

strf_today = datetime.date.today().strftime("%Y_%m_%d")

# Save the Excel file
wb.save(name + "_" + strf_today + '.xlsx')
# wb.remove_sheet()
print("done")


