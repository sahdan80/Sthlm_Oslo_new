#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Author: Daniel Sahlgren 2023, WSP

import emme_functions.emme as emme
import emme_functions.emme_functions as ef
import emme_functions.auxiliary_functions as aux
import pandas as pd
import inro
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
today = datetime.date.today()
# proj_name = "Stockolm-Oslo_255_UA1"
proj_name = "Stockolm-Oslo_255_JA"
# proj_name = "UA_ny"
scen_id = 1001 #scenario with time on us1
project_file_path = "D:/10350700_255_AE_DS/UA2/E444bank/JA/NB/Jvg/Jvg.emp"
# project_file_path = "R:/7055/10350700/5_Berakningar/Sampers/UA1_230421/E444bank/UA/NB/Jvg/Jvg.emp"
# ua_ver = "UA1"
# ua_ver = "UA2"
ua_ver = "JA"
my_emme = emme.Emme(project_file=project_file_path)
my_emmebank = my_emme.my_emmebank
my_emme.my_desktop.version
scenario = my_emmebank.scenario(scen_id)
network = scenario.get_network()

# get file with paths o get runt times from
strackor_file = "indata/strackor_restider.csv"
print(strackor_file)
strackor = pd.read_csv(strackor_file, encoding="UTF-8", sep=";")
strackor = [(x.strip(), y.strip()) for x, y in zip(strackor['start'], strackor['end'])]

if ua_ver =="UA2":# exclude Arvika-Lilleström
    my_excluded_links = [("5400","2310")] # links in i, j form
elif ua_ver=="UA1" or ua_ver=="JA": # exclude both alternatives of Gränsbanan
    my_excluded_links = [("5400","2310"), ("5400","5302")] # links in i, j form

my_excluded_links =[network.link(link_id[0], link_id[1]) for link_id in my_excluded_links] # get link objects
my_excluded_links_rev = [link.reverse_link for link in my_excluded_links] # get reverse links
my_excluded_links = my_excluded_links + my_excluded_links_rev
# dict w keys: statio name, va: node id
nodes = {}

for node in network.nodes():

    station = node["#station"]
    if len(station)>0: # if statoion name has been assigned to #station field in EMME
        nodes[station] = node.id
    else:
        pass

paths_to_csv = [] # listfor holding paths
my_paths = {}
excluded_links = []

for link in network.links():
    if not ef.check_link_mode(network=network, link=link, _modes=["i", "j", "k"]):
        excluded_links.append(link)

for stracka in strackor:
    try:
        nodes[stracka[0]]
    except KeyError:
        print(stracka[0] + "not in network. Consider revising the spelling of the station in input file")
    try:
        nodes[stracka[1]]
    except KeyError:
        print(stracka[1] + "not in network. Consider revising the spelling of the station in input file")

    start_id = nodes[stracka[0]]
    stop_id = nodes[stracka[1]]

    my_paths[stracka] = {}
    sp = network.shortest_path_tree(
        origin_node_id=start_id,
        link_costs='length',
        excluded_links=excluded_links+my_excluded_links,
        consider_turns=False,
        turn_costs=None,
        max_cost=None
    )

    # sp.reachable_nodes()
    path = sp.path_to_node(stop_id)
    exclude_lines = []
    for link in path:
        for seg in link.segments():
            if ef.is_transit(seg.line):
                line_itin = ef.get_itinerary_as_links(seg.line, _id=False, include_reverse=False, reverse_only=False)
                if all(item in line_itin for item in path): # check if the whole path is included in the itinerary
                    if seg.line.id not in my_paths[stracka]:
                        # my_paths[stracka][seg.line.id] = {"traveltime": seg.data1+seg.dwell_time, "trips": seg.line.data2}
                        my_paths[stracka][seg.line.id] = {"us1": seg.data1+seg.dwell_time, "traveltime": seg.data1+seg.dwell_time, "trips": seg.line.data2}
                    else:
                        my_paths[stracka][seg.line.id]["traveltime"] += (seg.data1+seg.dwell_time)  # sum total travel time on path
                        my_paths[stracka][seg.line.id]["us1"] += (seg.data1)  # sum total travel time on path
#
# # this part selects the line (original or reverse to include in selection)
# pop_lines = {} # holder for posob
# for str_, lines in my_paths.items(): # om lonjer vänder och kör tillbaka -> välj den med längt restid av retur resp originallinjen
#     pop_lines[str_] = []
#     for line_id, line_data in lines.items():
#         if line_id[-1]=="R":  # get last character if == R => reverse line
#             orig_line_id = line_id[:-1]
#             if orig_line_id in lines: #check original line:
#                 eval_list = [my_paths[str_][orig_line_id]["traveltime"], line_data["traveltime"]] # list with traveltime values from original and reverse line
#                 max_traveltime = max(eval_list) # get the longest run time from original and reverse line
#                 max_index = eval_list.index(max_traveltime) # get index of line with longest run time
#                 if max_index == 0: # meas that original line is relevant and reverse should pop
#                     pop_lines[str_].append(line_id)
#                 else:
#                     pop_lines[str_].append(orig_line_id)
#
# for stracka, popped_lines in pop_lines.items():
#     for line in popped_lines:
#         my_paths[stracka].pop(line)
from emme_functions.auxiliary_functions import convert_time_format as ct

for str_, lines in my_paths.items():
    for line_id, line_data in lines.items():
        paths_to_csv.append({"stracka": str_[0] + "-" + str_[1], "Linje": line_id, "restid inkl uppehåll": ct(round(line_data["traveltime"],0)), "restid utan uppehåll": ct(round(line_data["us1"],0)),
                             "antal turer": line_data["trips"]})





df = pd.DataFrame(paths_to_csv)


print("done")


#create Excel
wb = Workbook()
# ws = wb.active

# write all data to excelsheet

# Create a pivot table
run_times = wb.create_sheet('restider')
for r in dataframe_to_rows(df, index=True, header=True):
    run_times.append(r)


strf_today = datetime.date.today().strftime("%Y_%m_%d")

# Save the Excel file
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet']) # remove empty sheet
wb.save("restider_" + proj_name + "_" + strf_today + '.xlsx')
# wb.remove_sheet()
print("done")